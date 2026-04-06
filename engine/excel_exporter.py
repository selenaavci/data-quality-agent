"""Excel Export Engine - Generates styled Excel files with color-coded issues and risk levels."""
from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .issue_detector import Issue, ISSUE_COLORS
from .risk_scorer import RISK_COLORS, RISK_TEXT_COLORS


def export_to_excel(
    df: pd.DataFrame,
    issues: list[Issue],
    row_risks: dict[int, str],
) -> io.BytesIO:
    """Create a styled Excel workbook and return it as a BytesIO buffer."""
    wb = Workbook()

    # ── Sheet 1: Flagged Data ──────────────────────────────────────
    ws = wb.active
    ws.title = "Flagged Data"

    # Build a lookup: (row_idx, col) -> issue_type (keep highest priority)
    cell_issues: dict[tuple[int, str], str] = {}
    issue_priority = [
        "type_drift", "format_issue", "semantic_inconsistency",
        "range_violation", "missing_value", "duplicate",
        "sparse_column", "meaningless_feature",
    ]
    for issue in issues:
        if issue.row_idx is None:
            continue
        key = (issue.row_idx, issue.col)
        existing = cell_issues.get(key)
        if existing is None:
            cell_issues[key] = issue.issue_type
        else:
            if issue_priority.index(issue.issue_type) < issue_priority.index(existing):
                cell_issues[key] = issue.issue_type

    # Add Risk column to dataframe
    out_df = df.copy()
    out_df["__Risk__"] = ""
    for row_idx, risk in row_risks.items():
        if row_idx in out_df.index:
            out_df.at[row_idx, "__Risk__"] = risk

    # Write header
    headers = list(out_df.columns)
    headers = [h if h != "__Risk__" else "Risk" for h in headers]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    for excel_row, df_idx in enumerate(out_df.index, 2):
        for col_idx, col_name in enumerate(out_df.columns, 1):
            val = out_df.at[df_idx, col_name]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.border = thin_border

            actual_col = col_name if col_name != "__Risk__" else "Risk"

            # Color issue cells
            if col_name != "__Risk__":
                issue_key = (df_idx, col_name)
                if col_name == "__all__":
                    pass
                if issue_key in cell_issues:
                    color = ISSUE_COLORS[cell_issues[issue_key]]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                # Also color full-row issues (duplicates)
                dup_key = (df_idx, "__all__")
                if dup_key in cell_issues and issue_key not in cell_issues:
                    color = ISSUE_COLORS["duplicate"]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Color Risk column
            if col_name == "__Risk__" and val in RISK_COLORS:
                bg = RISK_COLORS[val]
                fg = RISK_TEXT_COLORS[val]
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font = Font(bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(ws.max_row + 1, 102)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # ── Sheet 2: Issue Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("Issue Summary")
    summary_headers = ["Issue Type", "Count", "Affected Columns"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")

    from .issue_detector import ISSUE_TYPES
    row = 2
    type_counts: dict[str, tuple[int, set]] = {}
    for issue in issues:
        if issue.issue_type not in type_counts:
            type_counts[issue.issue_type] = (0, set())
        cnt, cols = type_counts[issue.issue_type]
        type_counts[issue.issue_type] = (cnt + 1, cols | {issue.col})

    for itype, (count, cols) in type_counts.items():
        label = ISSUE_TYPES.get(itype, itype)
        color = ISSUE_COLORS.get(itype, "FFFFFF")
        ws2.cell(row=row, column=1, value=label).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        ws2.cell(row=row, column=2, value=count)
        ws2.cell(row=row, column=3, value=", ".join(sorted(cols)))
        row += 1

    for col_idx in range(1, 4):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 30

    # ── Sheet 3: Column-Level Issues ───────────────────────────────
    col_issues = [i for i in issues if i.row_idx is None]
    if col_issues:
        ws3 = wb.create_sheet("Column Issues")
        col_headers = ["Column", "Issue Type", "Detail"]
        for col_idx, h in enumerate(col_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        for r, issue in enumerate(col_issues, 2):
            color = ISSUE_COLORS.get(issue.issue_type, "FFFFFF")
            ws3.cell(row=r, column=1, value=issue.col)
            ws3.cell(row=r, column=2, value=ISSUE_TYPES.get(issue.issue_type, issue.issue_type)).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws3.cell(row=r, column=3, value=issue.detail)
        for col_idx in range(1, 4):
            ws3.column_dimensions[ws3.cell(row=1, column=col_idx).column_letter].width = 35

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
